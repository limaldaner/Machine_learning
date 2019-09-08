###########################################################################
#         ORB for crack and spalls feature extraction algorithm 
#
# Canada
# August 2019
#
# Oriented by: Dr. M. Shahbazi
# Author: Liége Maldaner
# E-mail: liege.malda@gmail.com
#
# Results:
###########################################################################

%matplotlib inline
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import cv2
import scipy
from scipy.misc import imread
import _pickle as pickle
import random
import os
from sklearn import svm
import csv
from os import listdir
from os.path import isfile, join
from openpyxl import load_workbook

#TRAINING DATA FEATURES EXTRACTION WITH ORB
DETERMINE_CLASS = int(1) # 0 for cracks and 1 for Spalls
training_data = np.array([])
training_data_labels = np.array([])

files_path = 'IMG_files_spalls'

#Number of images to be processed
list = os.listdir(files_path)
number_files = len(list)

#So the CSV feature line number is the same as the image name (also a number)
for key in np.arange(1, number_files, 1):
    category = DETERMINE_CLASS
    directory_path = files_path + '\\' + str(key) + '.jpg'
    #print(directory_path)
    
    # extract features!
    gray = cv2.imread(directory_path,0)
    #mask = cv2.imread(directory_path_mask,0)
    gray = cv2.resize(gray, (400, 250))  # resize so we're always comparing same-sized images
    #mask = cv2.resize(mask, (400, 250))  

    # Initiate ORB detector
    orb = cv2.ORB_create()

     # find the keypoints with ORB
    kp = orb.detect(gray,None)
    #ompute the descriptors with ORB
    kp1, des1 = orb.compute(gray, kp)

     # This is to make sure we have at least 100 keypoints to analyze
    # could also duplicate a few features if needed to hit a higher value
    if len(kp1) < 100:
        continue

    # transform the data to float and shuffle all keypoints
    # so we get a random sampling from each image
    des1 = des1.astype(np.float64)
    np.random.shuffle(des1)
    des1 = des1[0:100,:] # trim vector so all are same size
    vector_data = des1.reshape(1,3200) 
    list_data = vector_data.tolist()

    # We need to concatenate on the full list of features extracted from each image
    if len(training_data) == 0:
        training_data = np.append(training_data, vector_data)
        training_data = training_data.reshape(1,3200)
        training_data_labels = np.append(training_data_labels,np.append(vector_data,category))
    else:
        training_data   = np.concatenate((training_data, vector_data), axis=0)
        training_data_labels = np.append(training_data_labels,np.append(vector_data,category))

        
        
#SAVE FEATURES AND CLASSES TO AN EXEL FILE (Sheet1)
xlsFilename = r"C:\Users\Liege Maldaner\Desktop\ORB_test1.xlsx"
df = pd.DataFrame((training_data_labels.reshape(training_data.shape[0],training_data.shape[1]+1)))

book = load_workbook(xlsFilename)
writer = pd.ExcelWriter(xlsFilename, engine='openpyxl')
writer.book = book
writer.sheets = {ws.title: ws for ws in book.worksheets}

df.to_excel(writer,sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, index = False,header= False) 
    
writer.save()
